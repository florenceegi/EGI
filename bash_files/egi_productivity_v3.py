#!/usr/bin/env python3
"""
📊 EGI Commit Statistics to Excel Converter v3.0
================================================

Versione ULTRA migliorata con:
- Lines touched (non solo added)
- TAG-weighted commits (FIX/REFACTOR valgono di più)
- Day type classification (REFACTORING vs FEATURE_DEV)
- Cognitive load estimation
- Multi-dimensional productivity scoring
- Satisfaction tracking correlation

@author: Padmin D. Curtis OS3.0 for Fabio Cherici
@version: 3.0.0 (FlorenceEGI - Advanced Productivity Analytics)
@date: 2025-09-30
@purpose: Misurare VALORE REALE, non solo attività
"""

import subprocess
import pandas as pd
import re
from datetime import datetime, timedelta
import sys
import os
from pathlib import Path
import json

class EGICommitStatsExporterV3:
    """
    Versione 3.0 con sistema di misurazione multi-dimensionale
    che distingue tra tipi di lavoro e misura valore reale
    """

    # TAG weights: quanto vale ogni tipo di commit
    TAG_WEIGHTS = {
        'FEAT': 1.0,      # Standard - nuove features
        'FIX': 1.5,       # Fix vale di più (previene emergencies)
        'REFACTOR': 2.0,  # Refactoring è preziosissimo (debt repayment)
        'DOC': 0.8,       # Importante ma meno impattante
        'TEST': 1.2,      # Testing ha valore alto (quality assurance)
        'CHORE': 0.6,     # Maintenance routine
        'DEBUG': 1.3,     # Debugging è complesso
        'UNTAGGED': 0.5   # Penalità per commit non categorizzati
    }

    # Day type classification criteria
    DAY_TYPES = {
        'REFACTORING': {
            'description': 'Debt Repayment Day',
            'criteria': lambda stats: stats.get('REFACTOR', 0) > 20 or stats.get('FIX', 0) > 50,
            'multiplier': 1.5,  # Refactoring vale tantissimo
            'color': 'FFA500',  # Orange
            'icon': '🔧'
        },
        'BUG_FIXING': {
            'description': 'Bug Extermination Day',
            'criteria': lambda stats: stats.get('FIX', 0) > 40 and stats.get('FIX', 0) <= 50,
            'multiplier': 1.3,
            'color': 'FF6B6B',  # Red
            'icon': '🐛'
        },
        'FEATURE_DEV': {
            'description': 'Feature Building Day',
            'criteria': lambda stats: stats.get('FEAT', 0) > 50,
            'multiplier': 1.0,  # Standard
            'color': '4ECDC4',  # Teal
            'icon': '✨'
        },
        'TESTING': {
            'description': 'Quality Assurance Day',
            'criteria': lambda stats: stats.get('TEST', 0) > 40,
            'multiplier': 1.1,
            'color': '95E1D3',  # Light green
            'icon': '🧪'
        },
        'MAINTENANCE': {
            'description': 'Maintenance Day',
            'criteria': lambda stats: stats.get('CHORE', 0) > 40,
            'multiplier': 0.8,
            'color': 'C7CEEA',  # Light purple
            'icon': '🔨'
        },
        'MIXED': {
            'description': 'Mixed Activities',
            'criteria': lambda stats: True,  # Default fallback
            'multiplier': 1.0,
            'color': 'B8B8B8',  # Gray
            'icon': '📦'
        }
    }

    def __init__(self):
        self.git_repo_path = Path(__file__).parent.parent
        self.output_file = self.git_repo_path / "commit_statistics_v3.xlsx"
        self.tag_patterns = {
            'FEAT': r'\[FEAT\]',
            'FIX': r'\[FIX\]',
            'REFACTOR': r'\[REFACTOR\]',
            'DOC': r'\[DOC\]',
            'TEST': r'\[TEST\]',
            'CHORE': r'\[CHORE\]',
            'DEBUG': r'\[DEBUG\]'
        }

    def run_git_command(self, command):
        """Esegue un comando git e ritorna l'output"""
        try:
            result = subprocess.run(
                command,
                shell=True,
                cwd=self.git_repo_path,
                capture_output=True,
                text=True,
                check=True
            )
            return result.stdout.strip()
        except subprocess.CalledProcessError as e:
            print(f"Errore git command: {e}")
            return ""

    def get_commits_for_period(self, start_date, end_date):
        """Ottieni tutti i commit per un periodo specifico (con messaggi completi)"""
        cmd = f'git log --oneline --since="{start_date}" --until="{end_date} 23:59:59"'
        output = self.run_git_command(cmd)

        if not output:
            return []

        commits = []
        for line in output.split('\n'):
            if line.strip():
                commits.append(line.strip())

        return commits

    def calculate_weighted_commits(self, commits):
        """
        Calcola commit pesati basandosi sui TAG

        Un commit [REFACTOR] vale 2.0x
        Un commit [FIX] vale 1.5x
        Un commit [FEAT] vale 1.0x
        etc.
        """
        weighted_total = 0
        tag_breakdown = {tag: 0 for tag in self.tag_patterns.keys()}
        tag_breakdown['UNTAGGED'] = 0

        for commit in commits:
            weight = self.TAG_WEIGHTS['UNTAGGED']  # Default
            tag_found = False

            for tag, pattern in self.tag_patterns.items():
                if re.search(pattern, commit):
                    weight = self.TAG_WEIGHTS[tag]
                    tag_breakdown[tag] += 1
                    tag_found = True
                    break

            if not tag_found:
                tag_breakdown['UNTAGGED'] += 1

            weighted_total += weight

        return weighted_total, tag_breakdown

    def analyze_commits(self, commits):
        """Analizza i commit e categorizza per TAG"""
        stats = {
            'total_commits': len(commits),
            'tagged_commits': 0,
            'untagged_commits': 0,
            'tags': {tag: 0 for tag in self.tag_patterns.keys()}
        }

        for commit in commits:
            has_tag = False
            for tag, pattern in self.tag_patterns.items():
                if re.search(pattern, commit):
                    stats['tags'][tag] += 1
                    has_tag = True
                    break

            if has_tag:
                stats['tagged_commits'] += 1
            else:
                stats['untagged_commits'] += 1

        # Calcola percentuali
        if stats['total_commits'] > 0:
            stats['tag_coverage'] = round((stats['tagged_commits'] / stats['total_commits']) * 100, 1)
        else:
            stats['tag_coverage'] = 0

        return stats

    def classify_day_type(self, tag_stats):
        """
        Classifica il tipo di giornata basandosi sulla distribuzione TAG

        Returns: (day_type_name, day_config_dict)
        """
        total = sum(tag_stats.values())
        if total == 0:
            return 'MIXED', self.DAY_TYPES['MIXED']

        # Calcola percentuali
        percentages = {tag: (count/total)*100 for tag, count in tag_stats.items()}

        # Prova i criteri in ordine di priorità
        for day_type in ['REFACTORING', 'BUG_FIXING', 'FEATURE_DEV', 'TESTING', 'MAINTENANCE']:
            if self.DAY_TYPES[day_type]['criteria'](percentages):
                return day_type, self.DAY_TYPES[day_type]

        return 'MIXED', self.DAY_TYPES['MIXED']

    def estimate_cognitive_load(self, commits_count, files_modified, lines_touched, tag_stats):
        """
        Stima il cognitive load della giornata

        Fattori che aumentano cognitive load:
        - Context switching (molti file)
        - Lines touched elevate (tanto codice da gestire)
        - Molti FIX (cascata di problemi)
        - Refactoring profondo
        """
        base_load = 1.0

        # Context switching penalty (cambio file frequente)
        if files_modified > 15:
            base_load += (files_modified - 15) * 0.05

        # Lines touched intensity
        if lines_touched > 3000:
            base_load += (lines_touched - 3000) / 3000 * 0.5

        # Fix cascade penalty (ogni FIX è un problema scoperto)
        fix_count = tag_stats.get('FIX', 0)
        if fix_count > 10:
            base_load += (fix_count - 10) * 0.1

        # Refactoring depth (refactoring è cognitivamente pesante)
        refactor_count = tag_stats.get('REFACTOR', 0)
        if refactor_count > 3:
            base_load += refactor_count * 0.15

        # Debug complexity
        debug_count = tag_stats.get('DEBUG', 0)
        if debug_count > 0:
            base_load += debug_count * 0.2

        return round(base_load, 2)

    def get_code_lines_for_date(self, date_str):
        """
        Calcola le righe di codice per una data specifica
        MIGLIORAMENTO v3: Include 'touched' (added + removed)
        """
        today = datetime.now().date()
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()

        if target_date == today:
            cmd = f'git log --numstat --since="{date_str} 00:00:00" --pretty=format:""'
        else:
            next_date = datetime.strptime(date_str, '%Y-%m-%d') + timedelta(days=1)
            next_date_str = next_date.strftime('%Y-%m-%d')
            cmd = f'git log --numstat --since="{date_str} 00:00:00" --until="{next_date_str} 00:00:00" --pretty=format:""'

        output = self.run_git_command(cmd)

        if not output.strip():
            return {'added': 0, 'removed': 0, 'net': 0, 'touched': 0}

        total_added = 0
        total_removed = 0

        for line in output.split('\n'):
            if line.strip():
                parts = line.split('\t')
                if len(parts) >= 3:
                    # Skip .history folder files
                    filepath = parts[2] if len(parts) > 2 else ''
                    if filepath.startswith('.history/'):
                        continue
                
                if len(parts) >= 2:
                    try:
                        added = int(parts[0]) if parts[0] != '-' else 0
                        removed = int(parts[1]) if parts[1] != '-' else 0
                        total_added += added
                        total_removed += removed
                    except ValueError:
                        continue

        net_lines = total_added - total_removed
        touched_lines = total_added + total_removed  # NUOVO: righe toccate

        return {
            'added': total_added,
            'removed': total_removed,
            'net': net_lines,
            'touched': touched_lines  # NUOVA METRICA
        }

    def get_weekly_code_lines(self, start_date, end_date):
        """Calcola le righe di codice totali per una settimana (con touched)"""
        today = datetime.now().date()
        end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()

        if end_dt >= today:
            cmd = f'git log --numstat --since="{start_date} 00:00:00" --pretty=format:""'
        else:
            next_date = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
            next_date_str = next_date.strftime('%Y-%m-%d')
            cmd = f'git log --numstat --since="{start_date} 00:00:00" --until="{next_date_str} 00:00:00" --pretty=format:""'

        output = self.run_git_command(cmd)

        if not output.strip():
            return {'added': 0, 'removed': 0, 'net': 0, 'touched': 0}

        total_added = 0
        total_removed = 0

        for line in output.split('\n'):
            if line.strip():
                parts = line.split('\t')
                if len(parts) >= 3:
                    # Skip .history folder files
                    filepath = parts[2] if len(parts) > 2 else ''
                    if filepath.startswith('.history/'):
                        continue
                
                if len(parts) >= 2:
                    try:
                        added = int(parts[0]) if parts[0] != '-' else 0
                        removed = int(parts[1]) if parts[1] != '-' else 0
                        total_added += added
                        total_removed += removed
                    except ValueError:
                        continue

        net_lines = total_added - total_removed
        touched_lines = total_added + total_removed

        return {
            'added': total_added,
            'removed': total_removed,
            'net': net_lines,
            'touched': touched_lines
        }

    def calculate_productivity_index_v3(self, commits_list, lines_touched, total_productive_minutes,
                                       files_modified, tag_stats, alpha=0.4, beta=0.6):
        """
        VERSIONE 3.0: Multi-dimensional productivity index

        Formula:
        P = (Output Score × Quality Score × Day Type Multiplier × Cognitive Load) × 100

        Componenti:
        1. Output Score = α × (weighted_commits/hour) + β × (lines_touched/hour)
        2. Quality Score = basato su distribuzione TAG
        3. Day Type Multiplier = peso del tipo di giornata
        4. Cognitive Load = complessità del lavoro
        """
        if total_productive_minutes <= 0:
            return {
                'index': 0,
                'day_type': 'MIXED',
                'day_type_desc': 'Mixed Activities',
                'day_type_icon': '📦',
                'cognitive_load': 1.0,
                'components': {'output': 0, 'quality': 0, 'multiplier': 1.0},
                'weighted_commits': 0
            }

        total_hours = total_productive_minutes / 60

        # 1. Calcola weighted commits
        weighted_commits, _ = self.calculate_weighted_commits(commits_list)

        # 2. Output Score (usa lines_touched invece di lines_added)
        weighted_commits_per_hour = weighted_commits / total_hours
        lines_touched_per_hour = lines_touched / total_hours

        output_score = (alpha * weighted_commits_per_hour) + (beta * lines_touched_per_hour)

        # 3. Quality Score (basato su distribuzione TAG)
        quality_score = self.calculate_quality_score(tag_stats)

        # 4. Day Type Classification
        day_type, day_config = self.classify_day_type(tag_stats)
        multiplier = day_config['multiplier']

        # 5. Cognitive Load
        cognitive_load = self.estimate_cognitive_load(
            len(commits_list),
            files_modified,
            lines_touched,
            tag_stats
        )

        # Formula finale
        base_productivity = output_score * quality_score
        adjusted_productivity = base_productivity * multiplier * cognitive_load

        return {
            'index': round(adjusted_productivity * 100, 1),
            'day_type': day_type,
            'day_type_desc': day_config['description'],
            'day_type_icon': day_config['icon'],
            'cognitive_load': cognitive_load,
            'weighted_commits': round(weighted_commits, 1),
            'components': {
                'output': round(output_score, 2),
                'quality': round(quality_score, 2),
                'multiplier': multiplier,
                'cognitive_load': cognitive_load
            }
        }

    def calculate_quality_score(self, tag_stats):
        """
        Calcola quality score basato su distribuzione TAG

        - Alta % di REFACTOR/FIX = high quality (debt repayment)
        - Alta % di FEAT = standard quality
        - Alta % di CHORE = lower quality
        - Presenza di UNTAGGED = quality penalty
        """
        total = sum(tag_stats.values())
        if total == 0:
            return 1.0

        score = 1.0

        # Calcola percentuali
        percentages = {tag: (count/total)*100 for tag, count in tag_stats.items()}

        # Bonus per refactoring/fix (debt repayment)
        if percentages.get('REFACTOR', 0) > 20:
            score += 0.3
        if percentages.get('FIX', 0) > 30:
            score += 0.2

        # Bonus per testing
        if percentages.get('TEST', 0) > 20:
            score += 0.15

        # Penalty per untagged
        if percentages.get('UNTAGGED', 0) > 30:
            score -= 0.2

        # Bonus per documentation
        if percentages.get('DOC', 0) > 10:
            score += 0.1

        return max(0.5, min(score, 2.0))  # Clamp tra 0.5 e 2.0

    def get_files_modified_count(self, date_str):
        """Conta quanti file sono stati modificati in un giorno"""
        today = datetime.now().date()
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()

        if target_date == today:
            cmd = f'git log --name-only --since="{date_str} 00:00:00" --pretty=format:""'
        else:
            next_date = datetime.strptime(date_str, '%Y-%m-%d') + timedelta(days=1)
            next_date_str = next_date.strftime('%Y-%m-%d')
            cmd = f'git log --name-only --since="{date_str} 00:00:00" --until="{next_date_str} 00:00:00" --pretty=format:""'

        output = self.run_git_command(cmd)

        if not output.strip():
            return 0

        # Conta file unici
        files = set()
        for line in output.split('\n'):
            if line.strip():
                files.add(line.strip())

        return len(files)

    def get_testing_time_data(self, start_date, end_date):
        """Estrae dati di testing time dal log per il periodo specificato"""
        testing_log_path = self.git_repo_path / "storage" / "logs" / "testing_time.log"

        if not testing_log_path.exists():
            return {
                'total_minutes': 0,
                'sessions_count': 0,
                'avg_session_minutes': 0,
                'daily_breakdown': {}
            }

        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)

        total_minutes = 0
        sessions_count = 0
        daily_breakdown = {}

        try:
            with open(testing_log_path, 'r') as f:
                for line in f:
                    try:
                        data = json.loads(line.strip())
                        if data['action'] == 'TESTING_END':
                            timestamp = datetime.fromisoformat(data['timestamp'].replace('Z', '+00:00')).replace(tzinfo=None)

                            if start_dt <= timestamp < end_dt:
                                duration = abs(data.get('duration', 0))
                                date_key = timestamp.strftime('%Y-%m-%d')

                                total_minutes += duration
                                sessions_count += 1

                                if date_key not in daily_breakdown:
                                    daily_breakdown[date_key] = {'minutes': 0, 'sessions': 0}
                                daily_breakdown[date_key]['minutes'] += duration
                                daily_breakdown[date_key]['sessions'] += 1
                    except:
                        continue
        except:
            pass

        avg_session_minutes = total_minutes / sessions_count if sessions_count > 0 else 0

        return {
            'total_minutes': total_minutes,
            'sessions_count': sessions_count,
            'avg_session_minutes': round(avg_session_minutes, 1),
            'daily_breakdown': daily_breakdown
        }

    def get_daily_commits(self, start_date, end_date):
        """Ottieni commit giornalieri con tutte le metriche v3"""
        daily_stats = []
        current_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
        today = datetime.now().date()

        while current_date <= end_dt:
            date_str = current_date.strftime('%Y-%m-%d')

            if current_date.date() == today:
                cmd = f'git log --oneline --since="{date_str} 00:00:00"'
            else:
                next_date = current_date + timedelta(days=1)
                next_date_str = next_date.strftime('%Y-%m-%d')
                cmd = f'git log --oneline --since="{date_str} 00:00:00" --until="{next_date_str} 00:00:00"'

            output = self.run_git_command(cmd)
            commits = [line.strip() for line in output.split('\n') if line.strip()]
            commit_count = len(commits)

            # Code stats con touched
            code_stats = self.get_code_lines_for_date(date_str)

            # Files modified count
            files_modified = self.get_files_modified_count(date_str)

            daily_stats.append({
                'date': date_str,
                'day_name': current_date.strftime('%A'),
                'commits': commit_count,
                'commits_list': commits,  # Serve per weighted calculation
                'files_modified': files_modified,
                'lines_added': code_stats['added'],
                'lines_removed': code_stats['removed'],
                'lines_net': code_stats['net'],
                'lines_touched': code_stats['touched']  # NUOVO
            })

            current_date += timedelta(days=1)

        return daily_stats

    def generate_weekly_data(self):
        """Genera dati settimanali dal 19 agosto 2025 con metriche v3"""
        start_date = datetime.strptime('2025-08-19', '%Y-%m-%d')
        today = datetime.now()

        weeks = []
        week_number = 1
        current_monday = start_date

        week_descriptions = {
            1: 'Introduzione TAG system',
            2: 'Stabilizzazione',
            3: 'Consolidamento',
            4: 'Sviluppo avanzato',
            5: 'Ottimizzazione',
            6: 'Completamento features',
            7: 'Testing e refinement'
        }

        while current_monday <= today:
            current_sunday = current_monday + timedelta(days=6)
            week_end = min(current_sunday, today)

            month_names = {
                8: 'Agosto', 9: 'Settembre', 10: 'Ottobre', 11: 'Novembre', 12: 'Dicembre'
            }

            start_month = month_names.get(current_monday.month, current_monday.strftime('%B'))
            end_month = month_names.get(week_end.month, week_end.strftime('%B'))

            if current_monday.month == week_end.month:
                period = f"{current_monday.day}-{week_end.day} {start_month} 2025"
            else:
                period = f"{current_monday.day} {start_month} - {week_end.day} {end_month} 2025"

            description = week_descriptions.get(week_number, f'Sviluppo settimana {week_number}')
            if week_end.date() == today.date():
                description += ' (in corso)'

            weeks.append({
                'name': f'Settimana {week_number}',
                'period': period,
                'start_date': current_monday.strftime('%Y-%m-%d'),
                'end_date': week_end.strftime('%Y-%m-%d'),
                'description': description
            })

            current_monday += timedelta(days=7)
            week_number += 1

        weekly_data = []
        all_daily_data = []
        testing_summary = []

        for week in weeks:
            commits = self.get_commits_for_period(week['start_date'], week['end_date'])
            stats = self.analyze_commits(commits)
            testing_data = self.get_testing_time_data(week['start_date'], week['end_date'])
            code_stats = self.get_weekly_code_lines(week['start_date'], week['end_date'])

            estimated_coding_minutes = stats['total_commits'] * 22
            total_productive_minutes = testing_data['total_minutes'] + estimated_coding_minutes

            # Calcola weighted commits
            weighted_commits, _ = self.calculate_weighted_commits(commits)

            # Classifica tipo settimana
            day_type, day_config = self.classify_day_type(stats['tags'])

            # Stima files modified (approssimazione settimanale)
            files_estimate = stats['total_commits'] * 3  # Media ~3 file per commit

            # Cognitive load settimanale
            cognitive_load = self.estimate_cognitive_load(
                stats['total_commits'],
                files_estimate,
                code_stats['touched'],
                stats['tags']
            )

            # Productivity index v3
            productivity_result = self.calculate_productivity_index_v3(
                commits,
                code_stats['touched'],
                total_productive_minutes,
                files_estimate,
                stats['tags']
            )

            weekly_data.append({
                'Settimana': week['name'],
                'Periodo': week['period'],
                'Descrizione': week['description'],
                'Commit Totali': stats['total_commits'],
                'Commit Pesati': round(weighted_commits, 1),
                'Commit con TAG': stats['tagged_commits'],
                'Copertura TAG %': stats['tag_coverage'],
                'Righe Aggiunte': code_stats['added'],
                'Righe Rimosse': code_stats['removed'],
                'Righe Nette': code_stats['net'],
                'Righe Toccate': code_stats['touched'],
                'Tipo Settimana': f"{productivity_result['day_type_icon']} {productivity_result['day_type_desc']}",
                'Cognitive Load': cognitive_load,
                'Indice Produttività v3': productivity_result['index'],
                'Output Score': productivity_result['components']['output'],
                'Quality Score': productivity_result['components']['quality'],
                'Day Multiplier': productivity_result['components']['multiplier'],
                'FEAT': stats['tags']['FEAT'],
                'FIX': stats['tags']['FIX'],
                'REFACTOR': stats['tags']['REFACTOR'],
                'DOC': stats['tags']['DOC'],
                'TEST': stats['tags']['TEST'],
                'CHORE': stats['tags']['CHORE'],
                'Testing Minutes': testing_data['total_minutes'],
                'Coding Minutes (est)': estimated_coding_minutes,
                'Total Productive Minutes': total_productive_minutes
            })

            # Dati giornalieri
            daily_data = self.get_daily_commits(week['start_date'], week['end_date'])
            for day in daily_data:
                day['settimana'] = week['name']
                date_str = day['date']

                if date_str in testing_data['daily_breakdown']:
                    day['testing_minutes'] = testing_data['daily_breakdown'][date_str]['minutes']
                    day['testing_sessions'] = testing_data['daily_breakdown'][date_str]['sessions']
                else:
                    day['testing_minutes'] = 0
                    day['testing_sessions'] = 0

                day['coding_minutes_est'] = day['commits'] * 22
                day['total_productive_minutes'] = day['testing_minutes'] + day['coding_minutes_est']

                # Analizza TAG per questo giorno
                day_stats = self.analyze_commits(day['commits_list'])

                # Productivity v3 per giorno
                day_productivity = self.calculate_productivity_index_v3(
                    day['commits_list'],
                    day['lines_touched'],
                    day['total_productive_minutes'],
                    day['files_modified'],
                    day_stats['tags']
                )

                day['weighted_commits'] = day_productivity['weighted_commits']
                day['indice_produttivita_v3'] = day_productivity['index']

                # Safe access with fallback
                try:
                    day['day_type'] = f"{day_productivity['day_type_icon']} {day_productivity['day_type_desc']}"
                except KeyError as e:
                    print(f"🔍 DEBUG - Missing key: {e}")
                    print(f"🔍 Available keys: {list(day_productivity.keys())}")
                    day['day_type'] = f"📦 {day_productivity.get('day_type', 'Mixed Activities')}"

                day['cognitive_load'] = day_productivity['cognitive_load']

                # TAG breakdown
                for tag in ['FEAT', 'FIX', 'REFACTOR', 'DOC', 'TEST', 'CHORE']:
                    day[f'tag_{tag}'] = day_stats['tags'].get(tag, 0)

                all_daily_data.append(day)

            testing_summary.append({
                'Settimana': week['name'],
                'Periodo': week['period'],
                'Testing Totale (h)': round(testing_data['total_minutes'] / 60, 1),
                'Sessioni Totali': testing_data['sessions_count'],
                'Coding Stimato (h)': round(estimated_coding_minutes / 60, 1),
                'Tempo Produttivo (h)': round(total_productive_minutes / 60, 1),
                'Rapporto Testing/Coding': f"{round((testing_data['total_minutes'] / estimated_coding_minutes) * 100, 1)}%" if estimated_coding_minutes > 0 else "N/A"
            })

        return weekly_data, all_daily_data, testing_summary

    def create_excel_file(self):
        """Crea il file Excel con tutti i dati v3"""
        print("📊 Generazione statistiche AVANZATE v3.0...")

        weekly_data, daily_data, testing_summary = self.generate_weekly_data()

        df_weekly = pd.DataFrame(weekly_data)
        df_daily = pd.DataFrame(daily_data)
        df_testing = pd.DataFrame(testing_summary)

        # Summary con metriche v3
        total_commits = sum(week['Commit Totali'] for week in weekly_data)
        total_weighted = sum(week['Commit Pesati'] for week in weekly_data)
        total_tagged = sum(week['Commit con TAG'] for week in weekly_data)
        avg_coverage = round(sum(week['Copertura TAG %'] for week in weekly_data) / len(weekly_data), 1)
        total_testing_minutes = sum(week['Testing Minutes'] for week in weekly_data)
        total_coding_minutes = sum(week['Coding Minutes (est)'] for week in weekly_data)
        avg_productivity_v3 = round(sum(week['Indice Produttività v3'] for week in weekly_data) / len(weekly_data), 1)
        total_lines_touched = sum(week['Righe Toccate'] for week in weekly_data)

        summary_data = [{
            'Metrica': 'Commit Totali',
            'Valore': total_commits,
            'Note': 'Dal 19 agosto 2025'
        }, {
            'Metrica': 'Commit Pesati (Weighted)',
            'Valore': round(total_weighted, 1),
            'Note': 'FIX/REFACTOR valgono di più'
        }, {
            'Metrica': 'Commit con TAG',
            'Valore': total_tagged,
            'Note': f'{round((total_tagged/total_commits)*100, 1)}% del totale'
        }, {
            'Metrica': 'Copertura TAG Media',
            'Valore': f'{avg_coverage}%',
            'Note': 'Media delle settimane'
        }, {
            'Metrica': 'Righe Toccate Totali',
            'Valore': f'{total_lines_touched:,}',
            'Note': 'Added + Removed (lavoro reale)'
        }, {
            'Metrica': 'Testing Time Totale',
            'Valore': f'{round(total_testing_minutes/60, 1)}h',
            'Note': f'{total_testing_minutes} minuti'
        }, {
            'Metrica': 'Coding Time Stimato',
            'Valore': f'{round(total_coding_minutes/60, 1)}h',
            'Note': '22 min per commit'
        }, {
            'Metrica': 'Indice Produttività v3 Medio',
            'Valore': avg_productivity_v3,
            'Note': 'Multi-dimensional scoring'
        }]

        df_summary = pd.DataFrame(summary_data)

        # Scrivi Excel
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            df_summary.to_excel(writer, sheet_name='Riepilogo v3', index=False)
            df_weekly.to_excel(writer, sheet_name='Statistiche Settimanali v3', index=False)
            df_testing.to_excel(writer, sheet_name='Testing Time Analysis', index=False)
            df_daily.to_excel(writer, sheet_name='Commit Giornalieri v3', index=False)

            self.format_excel_sheets(writer)

        print(f"✅ File Excel v3 creato: {self.output_file}")
        print(f"📁 Percorso: {self.output_file.absolute()}")
        print(f"\n📊 METRICHE v3.0:")
        print(f"   Commit totali: {total_commits}")
        print(f"   Commit pesati: {round(total_weighted, 1)}")
        print(f"   Righe toccate: {total_lines_touched:,}")
        print(f"   Indice Produttività v3: {avg_productivity_v3}")
        print(f"\n💡 MIGLIORAMENTI v3:")
        print(f"   ✅ Lines touched (non solo added)")
        print(f"   ✅ TAG-weighted commits")
        print(f"   ✅ Day type classification")
        print(f"   ✅ Cognitive load estimation")
        print(f"   ✅ Multi-dimensional scoring")

        return str(self.output_file.absolute())

    def format_excel_sheets(self, writer):
        """Formatta i fogli Excel con colori per day types"""
        from openpyxl.styles import Font, PatternFill, Alignment

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]

            # Header styling
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

    def print_daily_terminal_output(self):
        """Output terminale migliorato per la giornata corrente"""
        today = datetime.now().strftime('%Y-%m-%d')

        # Get today's data
        daily_data = self.get_daily_commits(today, today)
        if not daily_data:
            print("Nessun dato per oggi")
            return

        day = daily_data[0]

        # Analisi TAG
        day_stats = self.analyze_commits(day['commits_list'])

        # Testing data
        testing_data = self.get_testing_time_data(today, today)
        day['testing_minutes'] = testing_data['daily_breakdown'].get(today, {}).get('minutes', 0)
        day['coding_minutes_est'] = day['commits'] * 22
        day['total_productive_minutes'] = day['testing_minutes'] + day['coding_minutes_est']

        # Productivity v3
        productivity = self.calculate_productivity_index_v3(
            day['commits_list'],
            day['lines_touched'],
            day['total_productive_minutes'],
            day['files_modified'],
            day_stats['tags']
        )

        # Terminal output
        print("\n" + "="*70)
        print("📅 FLORENCE EGI - PRODUTTIVITÀ GIORNALIERA v3.0")
        print("="*70)
        print(f"📅 Data: {today}")
        print(f"👤 Autore: fabio cherici")
        print()
        print("📊 RISULTATI GIORNALIERI")
        print("━"*70)
        print(f"📝 Commit: {day['commits']} (weighted: {productivity['weighted_commits']})")
        print(f"📁 File modificati: {day['files_modified']}")
        print(f"➕ Righe aggiunte: {day['lines_added']:,}")
        print(f"➖ Righe rimosse: {day['lines_removed']:,}")
        print(f"🔄 Righe toccate: {day['lines_touched']:,}")
        print(f"🚀 RIGHE NETTE: {day['lines_net']:+,}")
        print()
        print(f"🎯 TIPO GIORNATA: {productivity['day_type_icon']} {productivity['day_type_desc']}")
        print(f"🧠 Cognitive Load: {productivity['cognitive_load']}x", end="")
        if productivity['cognitive_load'] > 2.0:
            print(" (ALTO - giornata impegnativa)")
        elif productivity['cognitive_load'] > 1.5:
            print(" (MEDIO-ALTO)")
        else:
            print(" (NORMALE)")

        print(f"⚡ Indice Produttività v3: {productivity['index']}", end="")
        if productivity['day_type'] == 'REFACTORING':
            print(" (ECCELLENTE per refactoring)")
        elif productivity['day_type'] == 'BUG_FIXING':
            print(" (ECCELLENTE per bug fixing)")
        else:
            print()

        print()
        print("📊 DISTRIBUZIONE TAG")
        print("━"*70)

        # Ordina TAG per count
        sorted_tags = sorted(day_stats['tags'].items(), key=lambda x: x[1], reverse=True)
        for tag, count in sorted_tags:
            if count > 0:
                percentage = (count / day['commits']) * 100
                weight_info = f"(peso {self.TAG_WEIGHTS.get(tag, 1.0)}x)"
                print(f"[{tag}]: {count} ({percentage:.1f}%) {weight_info}")

        if day_stats['tags']['FIX'] > 10:
            print("\n⚠️ DEBT REPAYMENT DAY - Alta concentrazione di fix")

        print()
        print("💡 INSIGHTS")
        print("━"*70)

        # Insights basati su metriche
        if productivity['day_type'] == 'REFACTORING':
            print("✅ Alto valore creato (debt elimination)")
            print("✅ Production-readiness migliorata")

        if productivity['cognitive_load'] > 2.0:
            print("⚠️ Cognitive load elevato (considera break domani)")

        if day['lines_touched'] > 5000:
            print(f"📈 Volume codice elevato ({day['lines_touched']:,} righe toccate)")

        if day_stats['tags']['FIX'] > 10:
            estimated_prevented = day_stats['tags']['FIX'] // 2
            print(f"📈 Estimated emergency days prevented: {estimated_prevented}+")

        print()
        print("🎯 RACCOMANDAZIONI")
        print("━"*70)

        if productivity['day_type'] == 'REFACTORING' and productivity['cognitive_load'] > 2.0:
            print("• Domani: Feature development (ricompensa dopamina)")
            print("• Celebra i fix come vittorie separate")

        if day_stats['untagged_commits'] > 0:
            print(f"• {day_stats['untagged_commits']} commit senza TAG (usa TAG system)")

        print()
        print("━"*70)
        print(f"Generated by: Padmin D. Curtis OS3.0 | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("━"*70)


def main():
    """Funzione principale"""
    print("🚀 EGI Commit Statistics Excel Exporter v3.0")
    print("   Multi-Dimensional Productivity Analytics")
    print("="*70)

    exporter = EGICommitStatsExporterV3()

    try:
        # Terminal output per oggi
        exporter.print_daily_terminal_output()

        # Excel export
        print("\n📊 Generazione file Excel...")
        output_path = exporter.create_excel_file()

        print(f"\n🎉 Export completato con successo!")

        if os.path.exists(output_path):
            file_size = os.path.getsize(output_path)
            print(f"📁 Dimensione file: {file_size:,} bytes")

    except Exception as e:
        print(f"❌ Errore durante l'export: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
