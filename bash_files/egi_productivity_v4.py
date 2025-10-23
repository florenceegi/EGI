#!/usr/bin/env python3
"""
📊 EGI Commit Statistics to Excel Converter v3.1.2
==================================================

OPTIMIZED VERSION - Fix critical line counting bug + DEBUG

Versione ULTRA migliorata con:
- Lines touched (non solo added)
- TAG-weighted commits (FIX/REFACTOR valgono di più)
- Day type classification (REFACTORING vs FEATURE_DEV)
- Cognitive load estimation
- Multi-dimensional productivity scoring
- Satisfaction tracking correlation

🔧 v3.1 FIXES:
- Fixed parsing bug (separate if statements causing corruption)
- Complete filters for node_modules/, vendor/, lock files
- Excluded minified files, build directories
- Performance improvements

🔧 v3.1.1 FIXES:
- Excluded storage/testing/ (JSON data files with 260k+ lines)
- Excluded debug HTML files (albo_debug_*, etc.)
- Excluded deliberazioni_*.json (large datasets)
- Now counts ONLY real development work

🔧 v3.1.2 DEBUG:
- Added debug logging to see which files are excluded
- Shows skipped files count and lines
- Helps troubleshoot filter issues

@author: Padmin D. Curtis (AI Partner OS3.0) for Fabio Cherici
@version: 3.1.2 (FlorenceEGI - Advanced Productivity Analytics)
@date: 2025-10-23
@purpose: Misurare VALORE REALE, non solo attività (now with accurate line counts + debug)
"""

import subprocess
import pandas as pd
import re
from datetime import datetime, timedelta
import sys
import os
from pathlib import Path
import json
import math

class EGICommitStatsExporterV3:
    """
    Versione 3.1 con sistema di misurazione multi-dimensionale
    che distingue tra tipi di lavoro e misura valore reale

    FIX v3.1: Conteggio righe accurato escludendo dipendenze
    """

    # TAG weights: quanto vale ogni tipo di commit
    TAG_WEIGHTS = {
        'FEAT': 1.0,      # Standard - nuove features
        'FIX': 1.5,       # Fix vale di più (previene emergencies)
        'REFACTOR': 2.0,  # Refactoring è preziosissimo (debt repayment)
        'DOC': 0.8,       # Importante ma meno impattante
        'TEST': 1.2,      # Testing ha valore alto (quality assurance)
        'CHORE': 0.6,     # Maintenance routine
        'DEBUG': 1.3,     # Debugging è complesso
        'UNTAGGED': 0.5   # Penalità per commit non categorizzati
    }

    # Day type classification criteria
    DAY_TYPES = {
        'REFACTORING': {
            'description': 'Debt Repayment Day',
            'criteria': lambda stats: stats.get('REFACTOR', 0) > 20 or stats.get('FIX', 0) > 50,
            'multiplier': 1.5,  # Refactoring vale tantissimo
            'color': 'FFA500',  # Orange
            'icon': '🔧'
        },
        'BUG_FIXING': {
            'description': 'Bug Extermination Day',
            'criteria': lambda stats: stats.get('FIX', 0) > 40 and stats.get('FIX', 0) <= 50,
            'multiplier': 1.3,
            'color': 'FF6B6B',  # Red
            'icon': '🐛'
        },
        'FEATURE_DEV': {
            'description': 'Feature Building Day',
            'criteria': lambda stats: stats.get('FEAT', 0) > 50,
            'multiplier': 1.0,  # Standard
            'color': '4ECDC4',  # Teal
            'icon': '✨'
        },
        'TESTING': {
            'description': 'Quality Assurance Day',
            'criteria': lambda stats: stats.get('TEST', 0) > 40,
            'multiplier': 1.1,
            'color': '95E1D3',  # Light green
            'icon': '🧪'
        },
        'MAINTENANCE': {
            'description': 'Maintenance Day',
            'criteria': lambda stats: stats.get('CHORE', 0) > 40,
            'multiplier': 0.8,
            'color': 'C7CEEA',  # Light purple
            'icon': '🔨'
        },
        'MIXED': {
            'description': 'Mixed Activities',
            'criteria': lambda stats: True,  # Default fallback
            'multiplier': 1.0,
            'color': 'B8B8B8',  # Gray
            'icon': '📦'
        }
    }

    # 🔧 NEW v3.1: File patterns to exclude from line counting
    EXCLUDED_PATTERNS = [
        # Dependency directories
        'node_modules/',
        'vendor/',
        'bower_components/',

        # History and IDE
        '.history/',
        '.vscode/',
        '.idea/',

        # Build outputs
        'build/',
        'dist/',
        'out/',
        'coverage/',
        '.next/',

        # Lock files (huge and auto-generated)
        'package-lock.json',
        'composer.lock',
        'yarn.lock',
        'pnpm-lock.yaml',

        # Minified files
        '.min.js',
        '.min.css',
        '.bundle.js',

        # Other auto-generated
        '.map',
        '.cache/',
        'tmp/',
        'temp/',

        # Testing data files (JSON datasets, debug HTML)
        'storage/testing/',
        'deliberazioni_',  # Large JSON data files
        'albo_debug_',     # Debug HTML files
        'albo_structure_debug',
        'albo_real_page',

        # Data directories
        'storage/data/',
        'storage/logs/',
        'storage/dumps/',
    ]

    def __init__(self):
        self.git_repo_path = Path(__file__).parent.parent
        self.output_file = self.git_repo_path / "commit_statistics_v3.xlsx"
        self.tag_patterns = {
            'FEAT': r'\[FEAT\]',
            'FIX': r'\[FIX\]',
            'REFACTOR': r'\[REFACTOR\]',
            'DOC': r'\[DOC\]',
            'TEST': r'\[TEST\]',
            'CHORE': r'\[CHORE\]',
            'DEBUG': r'\[DEBUG\]'
        }

    def should_exclude_file(self, filepath):
        """
        🔧 NEW v3.1: Check if file should be excluded from line counting

        @param filepath: Path to check
        @return: True if file should be excluded
        """
        if not filepath:
            return True

        for pattern in self.EXCLUDED_PATTERNS:
            if pattern in filepath:
                return True

        return False

    def run_git_command(self, command):
        """Esegue un comando git e ritorna l'output"""
        try:
            result = subprocess.run(
                command,
                shell=True,
                cwd=self.git_repo_path,
                capture_output=True,
                text=True,
                check=True
            )
            return result.stdout.strip()
        except subprocess.CalledProcessError as e:
            print(f"Errore git command: {e}")
            return ""

    def get_commits_for_period(self, start_date, end_date):
        """Ottieni tutti i commit per un periodo specifico (con messaggi completi)"""
        cmd = f'git log --oneline --since="{start_date}" --until="{end_date} 23:59:59"'
        output = self.run_git_command(cmd)

        if not output:
            return []

        commits = []
        for line in output.split('\n'):
            if line.strip():
                commits.append(line.strip())

        return commits

    def calculate_weighted_commits(self, commits):
        """
        Calcola commit pesati basandosi sui TAG

        Un commit [REFACTOR] vale 2.0x
        Un commit [FIX] vale 1.5x
        Un commit [FEAT] vale 1.0x
        etc.
        """
        weighted_total = 0
        tag_breakdown = {tag: 0 for tag in self.tag_patterns.keys()}
        tag_breakdown['UNTAGGED'] = 0

        for commit in commits:
            weight = self.TAG_WEIGHTS['UNTAGGED']  # Default
            tag_found = False

            for tag, pattern in self.tag_patterns.items():
                if re.search(pattern, commit):
                    weight = self.TAG_WEIGHTS[tag]
                    tag_breakdown[tag] += 1
                    tag_found = True
                    break

            if not tag_found:
                tag_breakdown['UNTAGGED'] += 1

            weighted_total += weight

        return weighted_total, tag_breakdown

    def analyze_commits(self, commits):
        """Analizza i commit e categorizza per TAG"""
        stats = {
            'total_commits': len(commits),
            'tagged_commits': 0,
            'untagged_commits': 0,
            'tags': {tag: 0 for tag in self.tag_patterns.keys()}
        }

        for commit in commits:
            has_tag = False
            for tag, pattern in self.tag_patterns.items():
                if re.search(pattern, commit):
                    stats['tags'][tag] += 1
                    has_tag = True
                    break

            if has_tag:
                stats['tagged_commits'] += 1
            else:
                stats['untagged_commits'] += 1

        # Calcola percentuali
        if stats['total_commits'] > 0:
            stats['tag_coverage'] = round((stats['tagged_commits'] / stats['total_commits']) * 100, 1)
        else:
            stats['tag_coverage'] = 0

        return stats

    def classify_day_type(self, tag_stats):
        """
        Classifica il tipo di giornata basandosi sulla distribuzione TAG

        Returns: (day_type_name, day_config_dict)
        """
        total = sum(tag_stats.values())
        if total == 0:
            return 'MIXED', self.DAY_TYPES['MIXED']

        # Calcola percentuali
        percentages = {tag: (count/total)*100 for tag, count in tag_stats.items()}

        # Prova i criteri in ordine di priorità
        for day_type in ['REFACTORING', 'BUG_FIXING', 'FEATURE_DEV', 'TESTING', 'MAINTENANCE']:
            if self.DAY_TYPES[day_type]['criteria'](percentages):
                return day_type, self.DAY_TYPES[day_type]

        return 'MIXED', self.DAY_TYPES['MIXED']

    def estimate_cognitive_load(self, commits_count, files_modified, lines_touched, tag_stats):
        """
        v3.1.3 — Cognitive Load (log-scaled, normalized, clamped)
        Range: 1.0x (normale) → 3.5x (molto alto)
        - LI: impatto righe (log, satura ~60k)
        - FM: spread file (log, satura ~200)
        - DP: defect pressure = (FIX + 1.5*REFACTOR + 1.2*DEBUG) / commits
        """
        # 1) componenti in [0..1] con log-damping
        li = min(1.0, math.log1p(max(0, lines_touched) / 2000.0) / math.log1p(30.0))   # ~60k → 1.0
        fm = min(1.0, math.log1p(max(0, files_modified) / 10.0) / math.log1p(20.0))    # ~200 → 1.0

        fix = max(0, tag_stats.get('FIX', 0))
        refa = max(0, tag_stats.get('REFACTOR', 0))
        debug = max(0, tag_stats.get('DEBUG', 0))
        commits = max(1, int(commits_count))

        dp_raw = (fix + 1.5*refa + 1.2*debug) / commits
        dp = max(0.0, min(1.0, dp_raw))  # clamp

        # 2) pesi (somma ≤ 3.5)
        cl = 1.0 + 1.2*li + 0.9*fm + 0.7*dp

        # 3) hard clamp globale per evitare outlier
        return round(min(cl, 3.5), 2)

    def get_code_lines_for_date(self, date_str, debug=False):
        """
        🔧 FIXED v3.1.1: Calcola le righe di codice per una data specifica

        IMPROVEMENTS:
        - Fixed parsing bug (separate if statements)
        - Excludes node_modules/, vendor/, lock files
        - Accurate line counting for real work

        @param date_str: Date in YYYY-MM-DD format
        @param debug: If True, print debug info about excluded files
        @return: Dict with added, removed, net, touched lines
        """
        today = datetime.now().date()
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()

        if target_date == today:
            cmd = f'git log --numstat --since="{date_str} 00:00:00" --pretty=format:""'
        else:
            next_date = datetime.strptime(date_str, '%Y-%m-%d') + timedelta(days=1)
            next_date_str = next_date.strftime('%Y-%m-%d')
            cmd = f'git log --numstat --since="{date_str} 00:00:00" --until="{next_date_str} 00:00:00" --pretty=format:""'

        output = self.run_git_command(cmd)

        if not output.strip():
            return {'added': 0, 'removed': 0, 'net': 0, 'touched': 0}

        total_added = 0
        total_removed = 0
        skipped_files = 0
        skipped_lines = 0
        processed_files = 0

        # 🔧 DEBUG: Track top excluded files
        excluded_files_detail = []

        for line in output.split('\n'):
            if not line.strip():
                continue

            parts = line.split('\t')

            # 🔧 FIXED: Single unified check
            if len(parts) >= 3:
                filepath = parts[2]

                # Skip excluded files
                if self.should_exclude_file(filepath):
                    skipped_files += 1
                    try:
                        added = int(parts[0]) if parts[0] != '-' else 0
                        removed = int(parts[1]) if parts[1] != '-' else 0
                        lines_in_file = added + removed
                        skipped_lines += lines_in_file

                        if debug and lines_in_file > 1000:
                            excluded_files_detail.append((filepath, lines_in_file))
                    except:
                        pass
                    continue

                # Parse line counts
                try:
                    added = int(parts[0]) if parts[0] != '-' else 0
                    removed = int(parts[1]) if parts[1] != '-' else 0
                    total_added += added
                    total_removed += removed
                    processed_files += 1
                except ValueError:
                    # Binary file or corrupt data
                    continue

        net_lines = total_added - total_removed
        touched_lines = total_added + total_removed

        # 🔧 DEBUG output
        if debug:
            print(f"\n🔍 DEBUG LINE COUNTING:")
            print(f"   Files processed: {processed_files}")
            print(f"   Files skipped: {skipped_files}")
            print(f"   Lines skipped: {skipped_lines:,}")
            print(f"   Lines counted: {touched_lines:,}")
            if excluded_files_detail:
                print(f"\n   Top excluded files (>1000 lines):")
                for filepath, lines in sorted(excluded_files_detail, key=lambda x: x[1], reverse=True)[:10]:
                    print(f"      {lines:>8,} lines: {filepath}")

        return {
            'added': total_added,
            'removed': total_removed,
            'net': net_lines,
            'touched': touched_lines,
            'skipped_files': skipped_files,
            'skipped_lines': skipped_lines,
            'processed_files': processed_files
        }

    def get_weekly_code_lines(self, start_date, end_date):
        """
        🔧 FIXED v3.1: Calcola le righe di codice totali per una settimana

        Now excludes dependencies and auto-generated files
        """
        today = datetime.now().date()
        end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()

        if end_dt >= today:
            cmd = f'git log --numstat --since="{start_date} 00:00:00" --pretty=format:""'
        else:
            next_date = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
            next_date_str = next_date.strftime('%Y-%m-%d')
            cmd = f'git log --numstat --since="{start_date} 00:00:00" --until="{next_date_str} 00:00:00" --pretty=format:""'

        output = self.run_git_command(cmd)

        if not output.strip():
            return {'added': 0, 'removed': 0, 'net': 0, 'touched': 0}

        total_added = 0
        total_removed = 0
        skipped_files = 0

        for line in output.split('\n'):
            if not line.strip():
                continue

            parts = line.split('\t')

            # 🔧 FIXED: Single unified check
            if len(parts) >= 3:
                filepath = parts[2]

                # Skip excluded files
                if self.should_exclude_file(filepath):
                    skipped_files += 1
                    continue

                # Parse line counts
                try:
                    added = int(parts[0]) if parts[0] != '-' else 0
                    removed = int(parts[1]) if parts[1] != '-' else 0
                    total_added += added
                    total_removed += removed
                except ValueError:
                    continue

        net_lines = total_added - total_removed
        touched_lines = total_added + total_removed

        return {
            'added': total_added,
            'removed': total_removed,
            'net': net_lines,
            'touched': touched_lines,
            'skipped_files': skipped_files
        }

    def calculate_productivity_index_v3(self, commits_list, lines_touched, total_productive_minutes,
                                       files_modified, tag_stats, alpha=0.4, beta=0.6):
        """
        VERSIONE 3.0: Multi-dimensional productivity index

        Formula:
        P = (Output Score × Quality Score × Day Type Multiplier × Cognitive Load) × 100

        Componenti:
        1. Output Score = α × (weighted_commits/hour) + β × (lines_touched/hour)
        2. Quality Score = basato su distribuzione TAG
        3. Day Type Multiplier = peso del tipo di giornata
        4. Cognitive Load = complessità del lavoro
        """
        if total_productive_minutes <= 0:
            return {
                'index': 0,
                'day_type': 'MIXED',
                'day_type_desc': 'Mixed Activities',
                'day_type_icon': '📦',
                'cognitive_load': 1.0,
                'components': {'output': 0, 'quality': 0, 'multiplier': 1.0},
                'weighted_commits': 0
            }

        total_hours = total_productive_minutes / 60

        # 1. Calcola weighted commits
        weighted_commits, _ = self.calculate_weighted_commits(commits_list)

        # 2. Output Score (usa lines_touched invece di lines_added)
        weighted_commits_per_hour = weighted_commits / total_hours
        lines_touched_per_hour = lines_touched / total_hours

        output_score = (alpha * weighted_commits_per_hour) + (beta * lines_touched_per_hour)

        # 3. Quality Score (basato su distribuzione TAG)
        quality_score = self.calculate_quality_score(tag_stats)

        # 4. Day Type Classification
        day_type, day_config = self.classify_day_type(tag_stats)
        multiplier = day_config['multiplier']

        # 5. Cognitive Load
        cognitive_load = self.estimate_cognitive_load(
            len(commits_list),
            files_modified,
            lines_touched,
            tag_stats
        )

        # Final index
        productivity_index = output_score * quality_score * multiplier * cognitive_load

        return {
            'index': round(productivity_index, 2),
            'day_type': day_type,
            'day_type_desc': day_config['description'],
            'day_type_icon': day_config['icon'],
            'cognitive_load': cognitive_load,
            'components': {
                'output': round(output_score, 2),
                'quality': round(quality_score, 2),
                'multiplier': multiplier
            },
            'weighted_commits': round(weighted_commits, 1)
        }

    def calculate_quality_score(self, tag_stats):
        """
        Calcola quality score basato sulla distribuzione TAG

        High quality = More REFACTOR, TEST, FIX
        Low quality = More UNTAGGED, CHORE
        """
        total = sum(tag_stats.values())
        if total == 0:
            return 1.0

        percentages = {tag: (count/total)*100 for tag, count in tag_stats.items()}

        score = 1.0

        # Bonus per REFACTOR (molto prezioso)
        if percentages.get('REFACTOR', 0) > 20:
            score += 0.3

        # Bonus per TEST
        if percentages.get('TEST', 0) > 15:
            score += 0.2

        # Bonus per FIX (previene emergencies)
        if percentages.get('FIX', 0) > 20:
            score += 0.15

        # Penalty per untagged
        if percentages.get('UNTAGGED', 0) > 30:
            score -= 0.2

        # Bonus per documentation
        if percentages.get('DOC', 0) > 10:
            score += 0.1

        return max(0.5, min(score, 2.0))  # Clamp tra 0.5 e 2.0

    def get_files_modified_count(self, date_str):
        """Conta quanti file sono stati modificati in un giorno"""
        today = datetime.now().date()
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()

        if target_date == today:
            cmd = f'git log --name-only --since="{date_str} 00:00:00" --pretty=format:""'
        else:
            next_date = datetime.strptime(date_str, '%Y-%m-%d') + timedelta(days=1)
            next_date_str = next_date.strftime('%Y-%m-%d')
            cmd = f'git log --name-only --since="{date_str} 00:00:00" --until="{next_date_str} 00:00:00" --pretty=format:""'

        output = self.run_git_command(cmd)

        if not output.strip():
            return 0

        # Conta file unici (esclusi quelli da ignorare)
        files = set()
        for line in output.split('\n'):
            filepath = line.strip()
            if filepath and not self.should_exclude_file(filepath):
                files.add(filepath)

        return len(files)

    def get_testing_time_data(self, start_date, end_date):
        """Estrae dati di testing time dal log per il periodo specificato"""
        testing_log_path = self.git_repo_path / "storage" / "logs" / "testing_time.log"

        if not testing_log_path.exists():
            return {
                'total_minutes': 0,
                'sessions_count': 0,
                'avg_session_minutes': 0,
                'daily_breakdown': {}
            }

        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)

        total_minutes = 0
        sessions_count = 0
        daily_breakdown = {}

        try:
            with open(testing_log_path, 'r') as f:
                for line in f:
                    try:
                        data = json.loads(line.strip())
                        if data['action'] == 'TESTING_END':
                            timestamp = datetime.fromisoformat(data['timestamp'].replace('Z', '+00:00')).replace(tzinfo=None)

                            if start_dt <= timestamp < end_dt:
                                duration = abs(data.get('duration', 0))
                                date_key = timestamp.strftime('%Y-%m-%d')

                                total_minutes += duration
                                sessions_count += 1

                                if date_key not in daily_breakdown:
                                    daily_breakdown[date_key] = {'minutes': 0, 'sessions': 0}
                                daily_breakdown[date_key]['minutes'] += duration
                                daily_breakdown[date_key]['sessions'] += 1
                    except:
                        continue
        except:
            pass

        avg_session_minutes = total_minutes / sessions_count if sessions_count > 0 else 0

        return {
            'total_minutes': total_minutes,
            'sessions_count': sessions_count,
            'avg_session_minutes': round(avg_session_minutes, 1),
            'daily_breakdown': daily_breakdown
        }

    def get_daily_commits(self, start_date, end_date):
        """Ottieni commit giornalieri con tutte le metriche v3"""
        daily_stats = []
        current_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
        today = datetime.now().date()

        while current_date <= end_dt:
            date_str = current_date.strftime('%Y-%m-%d')

            if current_date.date() == today:
                cmd = f'git log --oneline --since="{date_str} 00:00:00"'
            else:
                next_date = current_date + timedelta(days=1)
                next_date_str = next_date.strftime('%Y-%m-%d')
                cmd = f'git log --oneline --since="{date_str} 00:00:00" --until="{next_date_str} 00:00:00"'

            output = self.run_git_command(cmd)
            commits = [line.strip() for line in output.split('\n') if line.strip()]
            commit_count = len(commits)

            # Code stats con touched
            code_stats = self.get_code_lines_for_date(date_str)

            # Files modified count
            files_modified = self.get_files_modified_count(date_str)

            daily_stats.append({
                'date': date_str,
                'day_name': current_date.strftime('%A'),
                'commits': commit_count,
                'commits_list': commits,  # Serve per weighted calculation
                'files_modified': files_modified,
                'lines_added': code_stats['added'],
                'lines_removed': code_stats['removed'],
                'lines_net': code_stats['net'],
                'lines_touched': code_stats['touched']  # NUOVO
            })

            current_date += timedelta(days=1)

        return daily_stats

    def generate_weekly_data(self):
        """Genera dati settimanali dal 19 agosto 2025 con metriche v3"""
        start_date = datetime.strptime('2025-08-19', '%Y-%m-%d')
        today = datetime.now()

        weeks = []
        week_number = 1
        current_monday = start_date

        week_descriptions = {
            1: 'Introduzione TAG system',
            2: 'Stabilizzazione',
            3: 'Consolidamento',
            4: 'Sviluppo avanzato',
            5: 'Ottimizzazione',
            6: 'Completamento features',
            7: 'Testing e refinement',
            8: 'Advanced development',
            9: 'Production readiness'
        }

        while current_monday <= today:
            current_sunday = current_monday + timedelta(days=6)
            week_end = min(current_sunday, today)

            month_names = {
                8: 'Agosto', 9: 'Settembre', 10: 'Ottobre', 11: 'Novembre', 12: 'Dicembre'
            }

            start_month = month_names.get(current_monday.month, current_monday.strftime('%B'))
            end_month = month_names.get(week_end.month, week_end.strftime('%B'))

            if current_monday.month == week_end.month:
                period = f"{current_monday.day}-{week_end.day} {start_month} 2025"
            else:
                period = f"{current_monday.day} {start_month} - {week_end.day} {end_month} 2025"

            description = week_descriptions.get(week_number, f'Sviluppo settimana {week_number}')
            if week_end.date() == today.date():
                description += ' (in corso)'

            weeks.append({
                'name': f'Settimana {week_number}',
                'period': period,
                'start_date': current_monday.strftime('%Y-%m-%d'),
                'end_date': week_end.strftime('%Y-%m-%d'),
                'description': description
            })

            current_monday += timedelta(days=7)
            week_number += 1

        weekly_data = []
        all_daily_data = []
        testing_summary = []

        for week in weeks:
            commits = self.get_commits_for_period(week['start_date'], week['end_date'])
            stats = self.analyze_commits(commits)
            testing_data = self.get_testing_time_data(week['start_date'], week['end_date'])
            code_stats = self.get_weekly_code_lines(week['start_date'], week['end_date'])

            estimated_coding_minutes = stats['total_commits'] * 22
            total_productive_minutes = testing_data['total_minutes'] + estimated_coding_minutes

            # Calcola weighted commits
            weighted_commits, _ = self.calculate_weighted_commits(commits)

            # Classifica tipo settimana
            day_type, day_config = self.classify_day_type(stats['tags'])

            # Get daily data for this week
            daily_data = self.get_daily_commits(week['start_date'], week['end_date'])

            # Calcola productivity v3 per ogni giorno
            for day in daily_data:
                day_stats = self.analyze_commits(day['commits_list'])
                testing_day = self.get_testing_time_data(day['date'], day['date'])

                day['testing_minutes'] = testing_day['daily_breakdown'].get(day['date'], {}).get('minutes', 0)
                day['coding_minutes_est'] = day['commits'] * 22
                day['total_productive_minutes'] = day['testing_minutes'] + day['coding_minutes_est']

                productivity = self.calculate_productivity_index_v3(
                    day['commits_list'],
                    day['lines_touched'],
                    day['total_productive_minutes'],
                    day['files_modified'],
                    day_stats['tags']
                )

                day['productivity_v3'] = productivity['index']
                day['day_type'] = productivity['day_type_icon'] + ' ' + productivity['day_type_desc']
                day['cognitive_load'] = productivity['cognitive_load']
                day['weighted_commits'] = productivity['weighted_commits']

                # Aggiungi dettagli TAG
                for tag in self.tag_patterns.keys():
                    day[f'tag_{tag.lower()}'] = day_stats['tags'].get(tag, 0)
                day['tag_untagged'] = day_stats.get('untagged_commits', 0)
                day['tag_coverage_%'] = day_stats.get('tag_coverage', 0)

            all_daily_data.extend(daily_data)

            # Testing summary per settimana
            testing_summary.append({
                'week': week['name'],
                'period': week['period'],
                'total_minutes': testing_data['total_minutes'],
                'sessions': testing_data['sessions_count'],
                'avg_session_minutes': testing_data['avg_session_minutes'],
                'estimated_coding_minutes': estimated_coding_minutes,
                'total_productive_minutes': total_productive_minutes,
                'coding_ratio_%': round((estimated_coding_minutes / total_productive_minutes * 100), 1) if total_productive_minutes > 0 else 0,
                'testing_ratio_%': round((testing_data['total_minutes'] / total_productive_minutes * 100), 1) if total_productive_minutes > 0 else 0
            })

            weekly_data.append({
                'week': week['name'],
                'period': week['period'],
                'description': week['description'],
                'commits': stats['total_commits'],
                'weighted_commits': round(weighted_commits, 1),
                'day_type': day_config['icon'] + ' ' + day_config['description'],
                'lines_added': code_stats['added'],
                'lines_removed': code_stats['removed'],
                'lines_net': code_stats['net'],
                'lines_touched': code_stats['touched'],
                'testing_minutes': testing_data['total_minutes'],
                'coding_minutes_est': estimated_coding_minutes,
                'total_productive_minutes': total_productive_minutes,
                'tag_coverage_%': stats['tag_coverage'],
                'tag_feat': stats['tags']['FEAT'],
                'tag_fix': stats['tags']['FIX'],
                'tag_refactor': stats['tags']['REFACTOR'],
                'tag_doc': stats['tags']['DOC'],
                'tag_test': stats['tags']['TEST'],
                'tag_chore': stats['tags']['CHORE'],
                'tag_debug': stats['tags']['DEBUG'],
                'tag_untagged': stats['untagged_commits']
            })

        return weekly_data, all_daily_data, testing_summary

    def create_excel_file(self):
        """Crea il file Excel con tutti i dati v3"""
        print("📊 Generazione dati settimanali e giornalieri...")
        weekly_data, daily_data, testing_data = self.generate_weekly_data()

        df_weekly = pd.DataFrame(weekly_data)
        df_daily = pd.DataFrame(daily_data)
        df_testing = pd.DataFrame(testing_data)

        # Calcola totali e medie per il riepilogo
        total_commits = df_weekly['commits'].sum()
        total_weighted = df_weekly['weighted_commits'].sum()
        total_lines_touched = df_weekly['lines_touched'].sum()
        total_testing_minutes = df_testing['total_minutes'].sum()
        total_coding_minutes = df_testing['estimated_coding_minutes'].sum()
        avg_coverage = df_weekly['tag_coverage_%'].mean()
        avg_productivity_v3 = df_daily['productivity_v3'].mean() if len(df_daily) > 0 else 0

        summary_data = [{
            'Metrica': 'Commit Totali',
            'Valore': total_commits,
            'Note': 'Tutti i commit nel periodo'
        }, {
            'Metrica': 'Commit Pesati Totali',
            'Valore': round(total_weighted, 1),
            'Note': 'Basato su TAG weights (REFACTOR=2x, FIX=1.5x)'
        }, {
            'Metrica': 'Copertura TAG Media',
            'Valore': f'{avg_coverage}%',
            'Note': 'Media delle settimane'
        }, {
            'Metrica': 'Righe Toccate Totali',
            'Valore': f'{total_lines_touched:,}',
            'Note': 'Added + Removed (real work, NO deps/testing data)'
        }, {
            'Metrica': 'Testing Time Totale',
            'Valore': f'{round(total_testing_minutes/60, 1)}h',
            'Note': f'{total_testing_minutes} minuti'
        }, {
            'Metrica': 'Coding Time Stimato',
            'Valore': f'{round(total_coding_minutes/60, 1)}h',
            'Note': '22 min per commit'
        }, {
            'Metrica': 'Indice Produttività v3 Medio',
            'Valore': round(avg_productivity_v3, 2),
            'Note': 'Multi-dimensional scoring'
        }]

        df_summary = pd.DataFrame(summary_data)

        # Scrivi Excel
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            df_summary.to_excel(writer, sheet_name='Riepilogo v3.1.2', index=False)
            df_weekly.to_excel(writer, sheet_name='Statistiche Settimanali v3', index=False)
            df_testing.to_excel(writer, sheet_name='Testing Time Analysis', index=False)
            df_daily.to_excel(writer, sheet_name='Commit Giornalieri v3', index=False)

            self.format_excel_sheets(writer)

        print(f"✅ File Excel v3.1.2 creato: {self.output_file}")
        print(f"📁 Percorso: {self.output_file.absolute()}")
        print(f"\n📊 METRICHE v3.1.2:")
        print(f"   Commit totali: {total_commits}")
        print(f"   Commit pesati: {round(total_weighted, 1)}")
        print(f"   Righe toccate: {total_lines_touched:,}")
        print(f"   Indice Produttività v3: {round(avg_productivity_v3, 2)}")
        print(f"\n🔧 FIXES v3.1.2:")
        print(f"   ✅ Fixed line counting bug")
        print(f"   ✅ Excluded node_modules/, vendor/")
        print(f"   ✅ Excluded storage/testing/ (JSON data)")
        print(f"   ✅ Excluded debug HTML files")
        print(f"   ✅ Debug logging enabled")
        print(f"   ✅ Accurate real work metrics")

        return str(self.output_file.absolute())

    def format_excel_sheets(self, writer):
        """Formatta i fogli Excel con colori per day types"""
        from openpyxl.styles import Font, PatternFill, Alignment

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]

            # Header styling
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

    def print_daily_terminal_output(self):
        """Output terminale migliorato per la giornata corrente"""
        today = datetime.now().strftime('%Y-%m-%d')

        # Get today's data
        daily_data = self.get_daily_commits(today, today)
        if not daily_data:
            print("Nessun dato per oggi")
            return

        day = daily_data[0]

        # Analisi TAG
        day_stats = self.analyze_commits(day['commits_list'])

        # 🔧 DEBUG: Re-run with debug enabled
        print("\n🔍 ANALISI ESCLUSIONI FILE...")
        code_stats_debug = self.get_code_lines_for_date(today, debug=True)

        # Testing data
        testing_data = self.get_testing_time_data(today, today)
        day['testing_minutes'] = testing_data['daily_breakdown'].get(today, {}).get('minutes', 0)
        day['coding_minutes_est'] = day['commits'] * 22
        day['total_productive_minutes'] = day['testing_minutes'] + day['coding_minutes_est']

        # Productivity v3
        productivity = self.calculate_productivity_index_v3(
            day['commits_list'],
            day['lines_touched'],
            day['total_productive_minutes'],
            day['files_modified'],
            day_stats['tags']
        )

        # Terminal output
        print("\n" + "="*70)
        print("📅 FLORENCE EGI - PRODUTTIVITÀ GIORNALIERA v3.1.2")
        print("="*70)
        print(f"📅 Data: {today}")
        print(f"👤 Autore: fabio cherici")
        print()
        print("📊 RISULTATI GIORNALIERI")
        print("━"*70)
        print(f"📝 Commit: {day['commits']} (weighted: {productivity['weighted_commits']})")
        print(f"📁 File modificati: {day['files_modified']}")
        print(f"➕ Righe aggiunte: {day['lines_added']:,}")
        print(f"➖ Righe rimosse: {day['lines_removed']:,}")
        print(f"🔄 Righe toccate: {day['lines_touched']:,} (escluse deps)")
        print(f"🚀 RIGHE NETTE: {day['lines_net']:+,}")
        print()
        print(f"🎯 TIPO GIORNATA: {productivity['day_type_icon']} {productivity['day_type_desc']}")
        print(f"🧠 Cognitive Load: {productivity['cognitive_load']}x", end="")
        if productivity['cognitive_load'] > 2.0:
            print(" (ALTO - giornata impegnativa)")
        elif productivity['cognitive_load'] > 1.5:
            print(" (MEDIO-ALTO)")
        else:
            print(" (NORMALE)")

        print(f"⚡ Indice Produttività v3: {productivity['index']}", end="")
        if productivity['day_type'] == 'REFACTORING':
            print(" (ECCELLENTE per refactoring)")
        elif productivity['day_type'] == 'BUG_FIXING':
            print(" (ECCELLENTE per bug fixing)")
        else:
            print()

        print()
        print("📊 DISTRIBUZIONE TAG")
        print("━"*70)

        # Ordina TAG per count
        sorted_tags = sorted(day_stats['tags'].items(), key=lambda x: x[1], reverse=True)
        for tag, count in sorted_tags:
            if count > 0:
                percentage = (count / day['commits']) * 100
                weight_info = f"(peso {self.TAG_WEIGHTS.get(tag, 1.0)}x)"
                print(f"[{tag}]: {count} ({percentage:.1f}%) {weight_info}")

        if day_stats['tags']['FIX'] > 10:
            print("\n⚠️ DEBT REPAYMENT DAY - Alta concentrazione di fix")

        print()
        print("💡 INSIGHTS")
        print("━"*70)

        # Insights basati su metriche
        if productivity['day_type'] == 'REFACTORING':
            print("✅ Alto valore creato (debt elimination)")
            print("✅ Production-readiness migliorata")

        if productivity['cognitive_load'] > 2.0:
            print("⚠️ Cognitive load elevato (considera break domani)")

        if day['lines_touched'] > 5000:
            print(f"📈 Volume codice elevato ({day['lines_touched']:,} righe toccate)")

        if day_stats['tags']['FIX'] > 10:
            estimated_prevented = day_stats['tags']['FIX'] // 2
            print(f"📈 Estimated emergency days prevented: {estimated_prevented}+")

        print()
        print("🎯 RACCOMANDAZIONI")
        print("━"*70)

        if productivity['day_type'] == 'REFACTORING' and productivity['cognitive_load'] > 2.0:
            print("• Domani: Feature development (ricompensa dopamina)")
            print("• Celebra i fix come vittorie separate")

        if day_stats['untagged_commits'] > 0:
            print(f"• {day_stats['untagged_commits']} commit senza TAG (usa TAG system)")

        print()
        print("━"*70)
        print(f"🔧 v3.1.2 DEBUG - See exclusions above | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("━"*70)


def main():
    """Funzione principale"""
    print("🚀 EGI Commit Statistics Excel Exporter v3.1.2 DEBUG")
    print("   Multi-Dimensional Productivity Analytics")
    print("   🔍 Shows which files are excluded from line counting")
    print("="*70)

    exporter = EGICommitStatsExporterV3()

    try:
        # Terminal output per oggi
        exporter.print_daily_terminal_output()

        # Excel export
        print("\n📊 Generazione file Excel...")
        output_path = exporter.create_excel_file()

        print(f"\n🎉 Export completato con successo!")

        if os.path.exists(output_path):
            file_size = os.path.getsize(output_path)
            print(f"📁 Dimensione file: {file_size:,} bytes")

    except Exception as e:
        print(f"❌ Errore durante l'export: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
