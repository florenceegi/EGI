"""
Add Excel charts following EXACT statistical guidelines from document
Rules:
1. Calculate proper vertical scale (min expected, max expected next 3 months)
2. Use straight lines (smooth=False)
3. Show numbers clearly on axes
4. Label weeks properly on X axis
5. One division = reasonable value (not automatic)
"""
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart.axis import TextAxis, NumericAxis
from pathlib import Path
import sys

def calculate_scale(data_values, log_name=""):
    """
    Calculate proper scale following 'Standard Admin' rules:
    The graph must clearly show the steepness of the trend (ups and downs).
    
    Rule 1: Never flatten the graph.
    Rule 2: Baseline does NOT need to be zero if data is high.
    Rule 3: Max should accommodate immediate future growth.
    
    Algorithm:
    1. Find true Min and Max of data.
    2. Calculate Range = Max - Min.
    3. If Range is 0 (all values same), force a range (e.g., +/- 10%).
    4. Set Min_Scale = Min - (Range * 0.10) -> Leave 10% breathing room at bottom.
    5. Set Max_Scale = Max + (Range * 0.20) -> Leave 20% breathing room at top for growth.
    6. Round Min/Max to 'nice' numbers (integers, 10s, 100s) for clean divisions.
    """
    if not data_values:
        return 0, 10
        
    true_min = min(data_values)
    true_max = max(data_values)
    
    # Handle single value or flatline
    if true_max == true_min:
        if true_max == 0:
            return 0, 10
        # Create artificial range around the value
        true_min = true_max * 0.9
        true_max = true_max * 1.1

    data_range = true_max - true_min
    
    # Padding
    target_min = true_min - (data_range * 0.15) # 15% clear below
    target_max = true_max + (data_range * 0.15) # 15% clear above
    
    # Ensure non-negative if data is all positive
    if true_min >= 0 and target_min < 0:
        target_min = 0

    # Rounding logic helper
    def round_nice(val, direction='down'):
        # Dynamic rounding based on magnitude
        magnitude = 1
        if abs(val) > 0:
            import math
            # Find order of magnitude (10, 100, 1000)
            magnitude = 10 ** math.floor(math.log10(abs(val)))
            # If value is like 550, magnitude is 100. nice step might be 50 or 100.
            if magnitude > 1:
                magnitude = magnitude / 10 # Refine granurality
                
        step = max(1, int(magnitude))
        
        if direction == 'down':
            return int(val / step) * step
        else:
            return int(val / step + 1) * step

    scale_min = round_nice(target_min, 'down')
    scale_max = round_nice(target_max, 'up')
    
    # Special Fix: If rounding caused the data to clip (rare but possible), fix it
    if scale_min > true_min: scale_min -= (scale_max - scale_min) / 10
    if scale_max < true_max: scale_max += (scale_max - scale_min) / 10

    print(f"      [Scaling {log_name}] Data: {true_min:.1f}-{true_max:.1f} | Range: {data_range:.1f} | Scale: {scale_min}-{scale_max}")
    return scale_min, scale_max

def add_charts_to_excel(excel_path: str):
    """Add Charts sheet with proper statistical graphs."""
    
    print(f"📊 Creating statistically correct charts for {excel_path}...")
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_path)
    
    if 'Weekly' not in wb.sheetnames:
        print("❌ No Weekly sheet found!")
        return False
    
    ws_weekly = wb['Weekly']
    num_weeks = ws_weekly.max_row - 1
    
    if num_weeks < 2:
        print("⚠️ Not enough data for charts")
        return False
    
    print(f"   Analyzing {num_weeks} weeks of data")
    
    # Remove old Charts sheet
    if 'Charts' in wb.sheetnames:
        del wb['Charts']
    
    ws_charts = wb.create_sheet('Charts', 0)
    
    # Title
    ws_charts['A1'] = 'GRAFICI PRODUTTIVITÀ - ANALISI STATISTICA SETTIMANALE'
    ws_charts['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_charts['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_charts['A1'].alignment = Alignment(horizontal='center')
    ws_charts.merge_cells('A1:J1')
    
    # Find columns
    header_row = list(ws_weekly[1])
    col_map = {}
    for idx, cell in enumerate(header_row, 1):
        if cell.value:
            col_map[cell.value] = idx
    
    # Extract data for scale calculation
    def get_column_data(col_name):
        if col_name not in col_map:
            return []
        col_idx = col_map[col_name]
        return [ws_weekly.cell(row, col_idx).value for row in range(2, num_weeks + 2) if ws_weekly.cell(row, col_idx).value]
    
    pi_data = get_column_data('Productivity Index')
    cp_data = get_column_data('Commits Pesati')
    rt_data = get_column_data('Righe Toccate')
    ct_data = get_column_data('Commits TOTALI')
    
    print(f"   Data ranges:")
    print(f"      Productivity Index: {min(pi_data) if pi_data else 0:.1f} - {max(pi_data) if pi_data else 0:.1f}")
    print(f"      Commits Pesati: {min(cp_data) if cp_data else 0:.1f} - {max(cp_data) if cp_data else 0:.1f}")
    print(f"      Righe Toccate: {min(rt_data) if rt_data else 0:,.0f} - {max(rt_data) if rt_data else 0:,.0f}")
    print(f"      Commits Totali: {min(ct_data) if ct_data else 0} - {max(ct_data) if ct_data else 0}")
    
        # Helper for consistent style
    def style_chart_admin(chart, title, y_title, x_title, values, category):
        chart.title = title
        chart.style = None # RESET PRESETS
        chart.width = 25
        chart.height = 15
        
        chart.y_axis.title = y_title
        chart.x_axis.title = x_title
        
        # SCALING
        min_scale, max_scale = calculate_scale(values, title)
        chart.y_axis.scaling.min = min_scale
        chart.y_axis.scaling.max = max_scale
        chart.y_axis.majorUnit = (max_scale - min_scale) / 10
        chart.y_axis.minorUnit = chart.y_axis.majorUnit / 5 # Detailed grid
        
        # GRIDLINES (Carta millimetrata simulation)
        from openpyxl.chart.axis import ChartLines
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.line import LineProperties
        from openpyxl.chart.label import DataLabelList
        
        chart.y_axis.majorGridlines = ChartLines()
        chart.y_axis.minorGridlines = ChartLines()
        chart.x_axis.majorGridlines = ChartLines()
        
        # AXIS LABELS
        chart.x_axis.tickLblPos = "low" # Ensure labels are at the bottom
        chart.x_axis.number_format = '@' # Text format for "Week 1", "Week 2"
        
        # ADD DATA
        chart.add_data(data, titles_from_data=True, from_rows=False)
        chart.set_categories(weeks)
        
        # SERIES STYLING (The core fix)
        # Access the created series and force styling
        if chart.series:
            s = chart.series[0]
            s.graphicalProperties.line.noFill = False
            s.graphicalProperties.line.solidFill = "000000" # Black line for contrast
            s.graphicalProperties.line.width = 20000 # 2pt
            s.smooth = False # FORCE STRAIGHT LINES
            
            # MARKERS (To show points clearly)
            s.marker.symbol = "circle"
            s.marker.graphicalProperties.solidFill = "FFFFFF" # White fill
            s.marker.graphicalProperties.line.solidFill = "000000" # Black outline

            # DATA LABELS (Values at nodes)
            s.dLbls = DataLabelList()
            s.dLbls.showVal = True
            s.dLbls.position = 't' # Top position to avoid overlapping line

        return chart

    # Chart 1: Productivity Index
    if 'Productivity Index' in col_map and pi_data:
        chart1 = LineChart()
        data = Reference(ws_weekly, min_col=col_map['Productivity Index'], min_row=1, max_row=num_weeks+1)
        weeks = Reference(ws_weekly, min_col=col_map['Settimana'], min_row=2, max_row=num_weeks+1)
        
        style_chart_admin(chart1, "Productivity Index (Settimanale)", "Score", "Settimana", pi_data, weeks)
        ws_charts.add_chart(chart1, "B2")
        print("   ✅ Chart 1: Productivity Index (CORRECTED: Straight lines + Markers)")
    
    # Chart 2: Commits Pesati
    if 'Commits Pesati' in col_map and cp_data:
        chart2 = LineChart()
        data = Reference(ws_weekly, min_col=col_map['Commits Pesati'], min_row=1, max_row=num_weeks+1)
        weeks = Reference(ws_weekly, min_col=col_map['Settimana'], min_row=2, max_row=num_weeks+1)
        
        style_chart_admin(chart2, "Commits Pesati (Settimanale)", "Weighted Value", "Settimana", cp_data, weeks)
        ws_charts.add_chart(chart2, "B32")
        print("   ✅ Chart 2: Commits Pesati (CORRECTED: Straight lines + Markers)")

    # Chart 3: Righe Toccate
    if 'Righe Toccate' in col_map and rt_data:
        chart3 = LineChart()
        data = Reference(ws_weekly, min_col=col_map['Righe Toccate'], min_row=1, max_row=num_weeks+1)
        weeks = Reference(ws_weekly, min_col=col_map['Settimana'], min_row=2, max_row=num_weeks+1)
        
        style_chart_admin(chart3, "Righe Toccate (Volume Lavoro)", "Lines", "Settimana", rt_data, weeks)
        ws_charts.add_chart(chart3, "N2") # Side by side
        print("   ✅ Chart 3: Righe Toccate (CORRECTED: Straight lines + Markers)")
    
    # Chart 4: Commits Totali
    if 'Commits TOTALI' in col_map and ct_data:
        chart4 = LineChart()
        data = Reference(ws_weekly, min_col=col_map['Commits TOTALI'], min_row=1, max_row=num_weeks+1)
        weeks = Reference(ws_weekly, min_col=col_map['Settimana'], min_row=2, max_row=num_weeks+1)
        
        style_chart_admin(chart4, "Commits Totali (Quantità)", "Count", "Settimana", ct_data, weeks)
        ws_charts.add_chart(chart4, "N32")
        print("   ✅ Chart 4: Commits Totali (CORRECTED: Straight lines + Markers)")
    
    # Save
    wb.save(excel_path)
    print(f"✅ Statistical charts created successfully")
    return True

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 add_excel_charts.py <excel_file>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    if not Path(excel_file).exists():
        print(f"❌ File not found: {excel_file}")
        sys.exit(1)
    
    success = add_charts_to_excel(excel_file)
    sys.exit(0 if success else 1)
